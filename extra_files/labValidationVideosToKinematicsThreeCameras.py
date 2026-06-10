"""
---------------------------------------------------------------------------
OpenCap: labValidationVideosToKinematics.py
---------------------------------------------------------------------------

Copyright 2022 Stanford University and the Authors

Author(s): Scott Uhlrich, Antoine Falisse

Licensed under the Apache License, Version 2.0 (the "License"); you may not
use this file except in compliance with the License. You may obtain a copy
of the License at http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

This script processes videos to estimate kinematics with the data from the 
"OpenCap: 3D human movement dynamics from smartphone videos" paper. The dataset
includes 10 subjects, with 2 sessions per subject. The first session includes
static, sit-to-stand, squat, and drop jump trials. The second session includes
walking trials.

The dataset is available on SimTK: https://simtk.org/frs/?group_id=2385. Make
sure you download the dataset with videos: LabValidation_withVideos.

In the paper, we compared 3 camera configurations: 
    2 cameras at +/- 45deg ('2-cameras'), 
    3 cameras at +/- 45deg and 0deg ('3-cameras'), and 
    5 cameras at +/- 45deg, +/- 70deg, and 0deg ('5-cameras'); 
where 0deg faces the participant. Use the variable cameraSetups below to select
which camera configuration to use. In the paper, we also compared three 
algorithms for pose detection: OpenPose at default resolution, OpenPose
with higher accuracy, and HRNet. HRNet is not supported on Windows, and we
therefore do not support it here (it is supported on the web application). To
use OpenPose at default resolution, set the variable resolutionPoseDetection to
'default'. To use OpenPose with higher accuracy, set the variable 
resolutionPoseDetection to '1x1008_4scales'. Take a look at 
Examples/reprocessSessions for more details about OpenPose settings and the
GPU requirements. Please note that we have updated OpenCap since submitting
the paper. As part of the updates, we re-trained the deep learning model we
use to predict anatomical markers from video keypoints, and we updated how
videos are time synchronized. These changes might have a slight effect
on the results.
"""

# %% Paths and imports.
import os
import sys
import shutil
import yaml

repoDir = os.path.abspath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)),'../'))
sys.path.append(repoDir)

from main import main
from utils import importMetadata

# Functions for excel update:
# %% Excel Progress Tracking Setup
PROGRESS_EXCEL_FILE = r'G:\Shared drives\Stanford Football\DataProcess.xlsx'

def update_progress_excel(subject_num, column_name, value):
    """
    Opens an Excel file, finds the correct cell, updates it, and saves the file.
    Finds the subject by Player ID in the first column, and updates the specified column.
    """
    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(PROGRESS_EXCEL_FILE)
        sheet = workbook.active

        # Find the target row by looking for the subject_num in the first column (Player ID)
        target_row = None
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=1).value
            # Handle both string and numeric comparisons
            if cell_value == subject_num or (isinstance(cell_value, (int, float)) and 
                                             isinstance(subject_num, (int, float)) and 
                                             int(cell_value) == int(subject_num)):
                target_row = row_index
                break
        
        if not target_row:
            print(f"  - EXCEL_UPDATE_WARNING: Subject ID '{subject_num}' not found in the first column of the Excel file.")
            return False

        # Find the target column by looking for the column_name in the first row
        target_col = None
        for col_index in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_index).value
            if cell_value and str(cell_value).lower() == str(column_name).lower():
                target_col = col_index
                break

        if not target_col:
            print(f"  - EXCEL_UPDATE_WARNING: Column '{column_name}' not found in the header row of the Excel file.")
            return False

        # Update the cell and save the workbook
        sheet.cell(row=target_row, column=target_col).value = value
        workbook.save(PROGRESS_EXCEL_FILE)
        print(f"  - EXCEL_UPDATE: Marked '{column_name}' as '{value}' for subject {subject_num}.")
        return True

    except FileNotFoundError:
        print(f"  - EXCEL_UPDATE_ERROR: The progress file was not found at '{PROGRESS_EXCEL_FILE}'.")
        return False
    except Exception as e:
        print(f"  - EXCEL_UPDATE_ERROR: An error occurred while updating the Excel file: {e}")
        return False


def extract_subject_id_from_session(session_name):
    """
    Extracts the subject ID number from a session name like 'subject17' -> 17
    """
    try:
        # Remove 'subject' prefix and convert to int
        if 'subject' in session_name.lower():
            subject_id_str = session_name.lower().replace('subject', '')
            return int(subject_id_str)
        else:
            # Try to extract number from the string
            numbers = re.findall(r'\d+', session_name)
            if numbers:
                return int(numbers[0])
    except (ValueError, AttributeError):
        pass
    return None

# %% User inputs
# Enter the path to the folder where you downloaded the data. The data is on
# SimTK: https://simtk.org/frs/?group_id=2385 (LabValidation_withVideos).
# In this example, our path looks like:
#   C:/Users/opencap/Documents/LabValidation_withVideos/subject2
#   C:/Users/opencap/Documents/LabValidation_withVideos/subject3
#   ...
dataDir = os.path.normpath('G:\Shared drives\Stanford Football\January_9')

# The dataset includes 2 sessions per subject.The first session includes
# static, sit-to-stand, squat, and drop jump trials. The second session 
# includes walking trials. The sessions are named <subject_name>_Session0 and 
# <subject_name>_Session1.
#need suvject 23, 28, 47 still, video synchronization errors
sessionNames = ['subject5']
#sessionNames = ['subject49','subject51','subject52','subject53','subject54','subject56','subject57','subject58','subject59','subject60','subject61','subject62','subject63','subject64','subject65','subject66','subject67','subject68']

#sessionNames = ['subject1','subject2','subject3','subject4','subject5', 'subject6', 'subject7', 'subject8', 'subject9', 'subject10','subject11', 'subject13', 'subject14', 'subject16', 'subject17','subject18', 'subject19','subject20','subject22','subject23', 'subject24','subject25','subject26','subject28','subject29','subject30','subject31','subject32','subject33','subject34','subject35','subject36','subject37','subject38','subject39','subject43','subject44','subject45','subject46','subject47','subject49','subject51','subject52','subject53','subject54','subject56','subject57','subject58','subject59','subject60','subject61','subject62','subject63','subject64','subject65','subject66','subject67','subject68']
# sessionNames = ['subject26','subject28','subject29','subject31','subject32','subject34','subject36','subject38','subject39','subject44','subject45','subject46','subject47','subject49','subject52','subject53','subject57','subject58','subject60','subject62','subject63','subject64','subject65','subject66','subject68']

# We only support OpenPose on Windows.
poseDetectors = ['OpenPose']

# Select the camera configuration you would like to use.
# cameraSetups = ['2-cameras', '3-cameras', '5-cameras']
cameraSetups = ['3-cameras']

# Select the resolution at which you would like to use OpenPose. More details
# about the options in Examples/reprocessSessions. In the paper, we compared 
# 'default' and '1x1008_4scales'.
resolutionPoseDetection = '1x1008_4scales'#'default' #'1x1008_4scales'   

# Since the prepint release, we updated a new augmenter model. To use the model
# used for generating the paper results, select v0.1. To use the latest model
# (now in production), select v0.2.
augmenter_model = 'v0.2'

# %% Data re-organization
# To reprocess the data, we need to re-organize the data so that the folder
# structure is the same one as the one expected by OpenCap. It is only done
# once as long as the variable overwriteRestructuring is False. To overwrite
# flip the flag to True.
overwriteRestructuring = False
subjects = sessionNames
for subject in subjects:
    pathSubject = os.path.join(dataDir, subject)
    pathVideos = os.path.join(pathSubject, 'Videos')    
    for session in os.listdir(pathVideos):
        if 'Session' not in session:
            continue
        pathSession = os.path.join(pathVideos, session)
        pathSessionNew = os.path.join(dataDir, 'Data', subject + '_' + session)
        if os.path.exists(pathSessionNew) and not overwriteRestructuring:
            continue
        os.makedirs(pathSessionNew, exist_ok=True)
        # Copy metadata
        pathMetadata = os.path.join(pathSubject, 'sessionMetadata.yaml')
        shutil.copy2(pathMetadata, pathSessionNew)
        pathMetadataNew = os.path.join(pathSessionNew, 'sessionMetadata.yaml')
        # Adjust model name
        sessionMetadata = importMetadata(pathMetadataNew)
        sessionMetadata['openSimModel'] = (
            'LaiUhlrich2022')
        with open(pathMetadataNew, 'w') as file:
                yaml.dump(sessionMetadata, file)        
        for cam in os.listdir(pathSession):
            if "Cam" not in cam:
                continue            
            pathCam = os.path.join(pathSession, cam)
            pathCamNew = os.path.join(pathSessionNew, 'Videos', cam)
            pathInputMediaNew = os.path.join(pathCamNew, 'InputMedia')
            # Copy videos.
            for trial in os.listdir(pathCam):
                pathTrial = os.path.join(pathCam, trial)
                if not os.path.isdir(pathTrial):
                    continue
                pathVideo = os.path.join(pathTrial, trial + '.mp4')
                pathTrialNew = os.path.join(pathInputMediaNew, trial)
                os.makedirs(pathTrialNew, exist_ok=True)
                shutil.copy2(pathVideo, pathTrialNew)
            # Copy camera parameters
            pathParameters = os.path.join(pathCam, 
                                          'cameraIntrinsicsExtrinsics.pickle')
            shutil.copy2(pathParameters, pathCamNew)

# %% Fixed settings.
# The dataset contains 5 videos per trial. The 5 videos are taken from cameras
# positioned at different angles: Cam0:-70deg, Cam1:-45deg, Cam2:0deg, 
# Cam3:45deg, and Cam4:70deg where 0deg faces the participant. Depending on the
# cameraSetup, we load different videos.
cam2sUse = {'5-cameras': ['Cam0', 'Cam1', 'Cam2', 'Cam3', 'Cam4'], 
            '3-cameras': ['Cam1', 'Cam4', 'Cam7'], 
            '2-cameras': ['Cam4', 'Cam5']}

# # %% Functions for re-processing the data.
def process_trial(trial_name=None, session_name=None, isDocker=False,
                  cam2Use=['all'],
                  intrinsicsFinalFolder='Deployed', extrinsicsTrial=False,
                  alternateExtrinsics=None, markerDataFolderNameSuffix=None,
                  imageUpsampleFactor=4, poseDetector='OpenPose',
                  resolutionPoseDetection='default', scaleModel=False,
                  bbox_thr=0.8, augmenter_model='v0.2', benchmark=False,
                  calibrationOptions=None, offset=True, dataDir=None):

    # Run main processing pipeline.
    main(session_name, trial_name, trial_name, cam2Use, intrinsicsFinalFolder,
          isDocker, extrinsicsTrial, alternateExtrinsics, calibrationOptions,
          markerDataFolderNameSuffix, imageUpsampleFactor, poseDetector,
          resolutionPoseDetection=resolutionPoseDetection,
          scaleModel=scaleModel, bbox_thr=bbox_thr,
          augmenter_model=augmenter_model, benchmark=benchmark, offset=offset,
          dataDir=dataDir)

    return

# %% Process trials.
for count, sessionName in enumerate(sessionNames):    
    # Get trial names.
    pathCam0 = os.path.join(dataDir, sessionName, 'Videos', 'Cam1',
                            'InputMedia')    
    # Work around to re-order trials and have the extrinsics trial firs, and
    # the static second (if available).
    trials_tmp = os.listdir(pathCam0)
    trials_tmp = [t for t in trials_tmp if
                  os.path.isdir(os.path.join(pathCam0, t))]
    session_with_static = True
    for trial in trials_tmp:
        if 'extrinsics' in trial.lower():                    
            extrinsics_idx = trials_tmp.index(trial) 
        if 'static' in trial.lower():                    
            static_idx = trials_tmp.index(trial) 
            session_with_static = True            
    trials = [trials_tmp[extrinsics_idx]]
    if session_with_static:
        trials.append(trials_tmp[static_idx])
        for trial in trials_tmp:
            if ('static' not in trial.lower() and 
                'extrinsics' not in trial.lower()):
                trials.append(trial)
    else:
        for trial in trials_tmp:
            if 'extrinsics' not in trial.lower():
                trials.append(trial)
    
    for poseDetector in poseDetectors:
        for cameraSetup in cameraSetups:
            cam2Use = cam2sUse[cameraSetup]
            
            # The second sessions (<>_1) have no static trial for scaling the
            # model. The static trials were collected as part of the first
            # session for each subject (<>_0). We here copy the Model folder
            # from the first session to the second session.
            if sessionName[-1] == '1':
                sessionDir = os.path.join(dataDir, sessionName)
                sessionDir_0 = sessionDir[:-1] + '1'
                camDir_0 = os.path.join(
                    sessionDir_0, 'OpenSimData', 
                    poseDetector + '_' + resolutionPoseDetection, cameraSetup)
                modelDir_0 = os.path.join(camDir_0, 'Model')
                camDir_1 = os.path.join(
                    sessionDir, 'OpenSimData', 
                    poseDetector + '_' + resolutionPoseDetection, cameraSetup)
                modelDir_1 = os.path.join(camDir_1, 'Model')
                os.makedirs(modelDir_1, exist_ok=True)
                for file in os.listdir(modelDir_0):
                    pathFile = os.path.join(modelDir_0, file)
                    pathFileEnd = os.path.join(modelDir_1, file)
                    shutil.copy2(pathFile, pathFileEnd)
                    
            # Process trial.
            for trial in trials:                
                print('Processing {}'.format(trial))
                
                # Detect if extrinsics trial to compute extrinsic parameters. 
                if 'extrinsics' in trial.lower():                    
                    extrinsicsTrial = True
                else:
                    extrinsicsTrial = False
                
                # Detect if static trial with neutral pose to scale model.
                if 'static' in trial.lower():                    
                    scaleModel = True
                else:
                    scaleModel = False
                
                # Session specific intrinsic parameters
                if 'subject222' in sessionName or 'subject333' in sessionName:
                    intrinsicsFinalFolder = 'Deployed_720_240fps'
                else:
                    intrinsicsFinalFolder = 'Deployed_720_60fps'                    
                    
                process_trial(trial,
                              session_name=sessionName,
                              cam2Use=cam2Use, 
                              intrinsicsFinalFolder=intrinsicsFinalFolder,
                              extrinsicsTrial=extrinsicsTrial,
                              markerDataFolderNameSuffix=cameraSetup,
                              poseDetector=poseDetector,
                              resolutionPoseDetection=resolutionPoseDetection,
                              scaleModel=scaleModel, 
                              augmenter_model=augmenter_model,
                              dataDir=dataDir)
                import openpyxl
                def update_progress_excel(column_name, value, sessionName):
                    workbook = openpyxl.load_workbook(PROGRESS_EXCEL_FILE)
                    subject_id = extract_subject_id_from_session(sessionName)
                    sheet = workbook.active# Find the target row by looking for the subject_num in the first column (Player ID)
                    target_row = None
                    for row_index in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row_index, column=1).value
                        # Handle both string and numeric comparisons
                        if cell_value == subject_id or (isinstance(cell_value, (int, float)) and 
                                                        isinstance(subject_id, (int, float)) and 
                                                        int(cell_value) == int(subject_id)):
                            target_row = row_index
                            break
                    # Find the target column by looking for the column_name in the first row
                    target_col = None
                    for col_index in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=1, column=col_index).value
                        if cell_value and str(cell_value).lower() == str(column_name).lower():
                            target_col = col_index
                            break
                    # Update the cell and save the workbook
                    sheet.cell(row=target_row, column=target_col).value = value
                    workbook.save(PROGRESS_EXCEL_FILE)

                update_progress_excel('neutral', '✓', sessionName)